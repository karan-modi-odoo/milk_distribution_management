import base64
import io
import logging

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Template column indices (0-based, matches the generated .xlsx template)
# ---------------------------------------------------------------------------
COL_DEALER = 0     # A — Dealer Name
COL_PRODUCT = 1    # B — Product Name
COL_QTY = 2        # C — Qty
COL_RATE = 3       # D — Rate (Rs)

# Row 1 (index 0) = header; Row 2 (index 1) = greyed example row
# Data starts at Row 3 (index 2)
HEADER_ROW_IDX = 0
EXAMPLE_ROW_IDX = 1
FIRST_DATA_ROW_IDX = 2


class MilkDispatchImportWizard(models.TransientModel):
    """
    Wizard — Import Dispatch Lines from Excel / ODS

    Allows the distributor to prepare dealer orders in a spreadsheet and
    import them directly into a **draft** Dispatch Sheet.

    Design constraints (strictly followed)
    ---------------------------------------
    * Works on DRAFT sheets only (server-side guard + UI button invisible
      on confirmed/delivered).
    * Zero existing model, field, workflow, or business logic is altered.
    * All writes go through the standard ORM so every existing constraint
      (_check_unique_partner_per_sheet, _check_unique_product_per_line,
      _check_qty) is automatically respected.
    * Existing lines on the sheet are never removed or overwritten.
    * A duplicate (dealer, product) combination is silently skipped and
      counted in the summary notification.
    * All validation errors are collected and surfaced together in one
      UserError so the user can fix the file in a single pass.
    * Rate is optional: blank → resolved from milk_rate_per_crate →
      lst_price, mirroring the onchange in MilkDispatchProductLine exactly.

    Supported file formats
    ----------------------
    * .xlsx  — parsed with openpyxl (always available in Odoo environments)
    * .ods   — parsed with pandas + odf engine (optional; graceful error if
               pandas/odf are not installed)
    """

    _name = 'milk.dispatch.import.wizard'
    _description = 'Import Dispatch Lines from Excel'

    sheet_id = fields.Many2one(
        'milk.dispatch.sheet',
        string='Dispatch Sheet',
        required=True,
        ondelete='cascade',
        readonly=True,
    )
    import_file = fields.Binary(
        string='Excel / ODS File',
        required=True,
        help=(
            'Upload the filled import template.\n'
            'Accepted: .xlsx (Excel 2007+) or .ods (LibreOffice Calc).\n'
            'Do NOT upload .xls (legacy Excel 97–2003).'
        ),
    )
    import_filename = fields.Char(string='File Name')

    # ------------------------------------------------------------------
    # Template download
    # ------------------------------------------------------------------

    def action_download_template(self):
        """
        Generate and stream a formatted .xlsx import template.

        Sheet 1 — 'Dispatch Import':
          Row 1 : bold, dark-blue header (frozen)
          Row 2 : italic, grey example row showing expected format
          Row 3+ : blank data entry area

        Sheet 2 — 'Instructions':
          Column-level documentation for each field.
        """
        self.ensure_one()
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError as exc:
            raise UserError(
                _(
                    'The "openpyxl" Python library is required to generate the '
                    'template.\n\nPlease ask your system administrator to run:\n'
                    '    pip install openpyxl'
                )
            ) from exc

        wb = openpyxl.Workbook()

        # ── Sheet 1: data entry ──────────────────────────────────────
        ws = wb.active
        ws.title = 'Dispatch Import'

        ws.column_dimensions['A'].width = 32
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 16

        thin = Side(style='thin', color='BFBFBF')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header row
        headers = ['Dealer Name', 'Product Name', 'Qty', 'Rate (Rs)']
        hdr_fill = PatternFill(
            start_color='1F4E79', end_color='1F4E79', fill_type='solid'
        )
        hdr_font = Font(bold=True, color='FFFFFF', size=11, name='Calibri')

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        ws.row_dimensions[1].height = 22

        # Example row (greyed, italic)
        example = ['Ramesh Dairy', 'Full Cream Milk 500ml', 10, 22.50]
        ex_fill = PatternFill(
            start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'
        )
        ex_font = Font(italic=True, color='595959', name='Calibri')
        for col_idx, val in enumerate(example, start=1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.fill = ex_fill
            cell.font = ex_font
            cell.alignment = Alignment(horizontal='left')
            cell.border = border

        # Freeze header + example rows
        ws.freeze_panes = 'A3'

        # ── Sheet 2: instructions ────────────────────────────────────
        ws2 = wb.create_sheet(title='Instructions')
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 40
        ws2.column_dimensions['C'].width = 58

        hdr2_font = Font(bold=True, name='Calibri')
        instructions = [
            ('Column', 'Description', 'Validation Rules'),
            (
                'Dealer Name',
                'Full name of the dealer contact in Odoo',
                'Must exactly match an existing res.partner name '
                '(case-insensitive). Required.',
            ),
            (
                'Product Name',
                'Full product display name as shown in Odoo',
                'Must exactly match product.product display_name '
                '(case-insensitive). Variant attributes are included. Required.',
            ),
            (
                'Qty',
                'Quantity in crates (or pieces for piece-based products)',
                'Must be a non-negative number. '
                'Half-crate and piece rules are validated by Odoo on import.',
            ),
            (
                'Rate (Rs)',
                'Billing rate per crate or per piece',
                'Optional. Leave blank to auto-fill from the product\'s '
                'configured default rate (milk_rate_per_crate → lst_price).',
            ),
        ]
        for row_num, row_data in enumerate(instructions, start=1):
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws2.cell(row=row_num, column=col_idx, value=val)
                if row_num == 1:
                    cell.font = hdr2_font

        # ── Serialize and create attachment ──────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'dispatch_import_template.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(buf.read()),
            'mimetype': (
                'application/vnd.openxmlformats-officedocument'
                '.spreadsheetml.sheet'
            ),
            'res_model': self._name,
            'res_id': self.id,
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    # ------------------------------------------------------------------
    # Import action
    # ------------------------------------------------------------------

    def action_import(self):
        """
        Parse the uploaded file, validate every row, and write the
        resulting dealer/product lines to the dispatch sheet.

        Returns a sticky display_notification with a full import summary.
        Raises UserError (with all validation errors joined) on failure.
        """
        self.ensure_one()
        sheet = self.sheet_id

        # ── Draft guard ───────────────────────────────────────────────
        if sheet.state != 'draft':
            raise UserError(
                _(
                    'Import is only allowed on Draft sheets.\n\n'
                    '"%s" is currently "%s". '
                    'You cannot import lines into a confirmed or delivered sheet.'
                )
                % (sheet.name, dict(sheet._fields['state'].selection).get(sheet.state, sheet.state))
            )

        if not self.import_file:
            raise UserError(_('Please upload an Excel (.xlsx) or ODS file first.'))

        # ── Parse ─────────────────────────────────────────────────────
        raw_rows = self._parse_file()

        if not raw_rows:
            raise UserError(
                _(
                    'No data rows were found in the uploaded file.\n\n'
                    'Please fill in the template starting from row 3 '
                    '(row 1 = header, row 2 = example).'
                )
            )

        # ── Validate ──────────────────────────────────────────────────
        import_plan, errors = self._validate_rows(raw_rows)

        if errors:
            raise UserError(
                _('The following errors were found in the import file. '
                  'Please fix them and upload again:\n\n%s')
                % '\n'.join(errors)
            )

        # ── Write ─────────────────────────────────────────────────────
        stats = self._apply_import_plan(import_plan, sheet)

        _logger.info(
            'Dispatch import — sheet: %s | new dealers: %d | '
            'new product lines: %d | skipped: %d',
            sheet.name,
            stats['new_dealers'],
            stats['new_products'],
            stats['skipped'],
        )

        msg = _(
            'Import completed successfully.\n'
            '• %(nd)d new dealer line(s) created\n'
            '• %(np)d new product line(s) added\n'
            '• %(sk)d row(s) skipped (already present on this sheet)'
        ) % {
            'nd': stats['new_dealers'],
            'np': stats['new_products'],
            'sk': stats['skipped'],
        }

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'type': 'success',
                'message': msg,
                'sticky': True,
            },
        }

    # ------------------------------------------------------------------
    # File parsing
    # ------------------------------------------------------------------

    def _parse_file(self):
        """
        Decode the uploaded binary and return a list of raw row dicts.

        Each dict has keys:
            row_num   (int, 1-based Excel row number for error messages)
            dealer    (str)
            product   (str)
            qty_raw   (raw cell value — float | int | str | None)
            rate_raw  (raw cell value — float | int | str | None)

        Fully blank rows (all four cells empty) are excluded.
        The header row and example row are always excluded.
        """
        filename = (self.import_filename or '').strip().lower()
        raw_bytes = base64.b64decode(self.import_file)

        if filename.endswith('.ods'):
            return self._parse_ods(raw_bytes)
        # Default: treat as .xlsx (covers files saved without extension too)
        return self._parse_xlsx(raw_bytes)

    # ── .xlsx parser ──────────────────────────────────────────────────

    def _parse_xlsx(self, raw_bytes):
        try:
            import openpyxl
        except ImportError as exc:
            raise UserError(
                _(
                    'The "openpyxl" Python library is required to parse .xlsx files.\n\n'
                    'Please ask your system administrator to run:\n'
                    '    pip install openpyxl'
                )
            ) from exc

        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(raw_bytes), read_only=True, data_only=True
            )
        except Exception as exc:
            raise UserError(
                _(
                    'Could not open the uploaded file as an Excel workbook.\n\n'
                    'Make sure you are uploading a valid .xlsx file (not .xls '
                    'or a renamed file).\n\nTechnical detail: %s'
                )
                % str(exc)
            ) from exc

        ws = wb.active
        rows = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx < FIRST_DATA_ROW_IDX:
                continue  # skip header + example
            excel_row_num = row_idx + 1  # 1-based

            dealer = self._to_str(row[COL_DEALER] if len(row) > COL_DEALER else None)
            product = self._to_str(row[COL_PRODUCT] if len(row) > COL_PRODUCT else None)
            qty_raw = row[COL_QTY] if len(row) > COL_QTY else None
            rate_raw = row[COL_RATE] if len(row) > COL_RATE else None

            # Skip fully blank rows silently
            if (
                not dealer
                and not product
                and qty_raw is None
                and rate_raw is None
            ):
                continue

            rows.append({
                'row_num': excel_row_num,
                'dealer': dealer,
                'product': product,
                'qty_raw': qty_raw,
                'rate_raw': rate_raw,
            })

        wb.close()
        return rows

    # ── .ods parser ───────────────────────────────────────────────────

    def _parse_ods(self, raw_bytes):
        try:
            import pandas as pd
        except ImportError as exc:
            raise UserError(
                _(
                    'The "pandas" Python library is required to import ODS files.\n\n'
                    'Please ask your system administrator to install it, or '
                    'save your file as .xlsx and upload that instead.'
                )
            ) from exc

        try:
            df = pd.read_excel(
                io.BytesIO(raw_bytes),
                engine='odf',
                header=HEADER_ROW_IDX,
                dtype=str,
            )
        except Exception as exc:
            raise UserError(
                _(
                    'Could not open the uploaded file as an ODS spreadsheet.\n\n'
                    'Technical detail: %s'
                )
                % str(exc)
            ) from exc

        rows = []
        for df_idx, row in df.iterrows():
            if df_idx == 0:
                continue  # skip the example row (index 0 after header skip)
            excel_row_num = df_idx + 2  # approximate display row

            vals = row.tolist()
            dealer = self._to_str(vals[COL_DEALER] if len(vals) > COL_DEALER else None)
            product = self._to_str(vals[COL_PRODUCT] if len(vals) > COL_PRODUCT else None)
            qty_raw = vals[COL_QTY] if len(vals) > COL_QTY else None
            rate_raw = vals[COL_RATE] if len(vals) > COL_RATE else None

            if not dealer and not product:
                continue

            rows.append({
                'row_num': excel_row_num,
                'dealer': dealer,
                'product': product,
                'qty_raw': qty_raw,
                'rate_raw': rate_raw,
            })

        return rows

    # ------------------------------------------------------------------
    # Validation
    # ------------------------------------------------------------------

    def _validate_rows(self, raw_rows):
        """
        Validate every raw row and build an import plan.

        Returns
        -------
        import_plan : list of dicts
            {row_num, partner_id, product_id, qty, rate (float|None),
             dealer_name, product_name}
        errors : list of str
            Human-readable error messages (empty list = no errors).
        """
        errors = []
        import_plan = []

        # Build lookup caches once for the whole import (avoids N+1 searches)
        partner_cache = self._build_partner_cache()
        product_cache = self._build_product_cache()

        for row in raw_rows:
            row_num = row['row_num']
            dealer_name = row['dealer']
            product_name = row['product']
            qty_raw = row['qty_raw']
            rate_raw = row['rate_raw']

            # ── Required fields ──────────────────────────────────────
            if not dealer_name:
                errors.append(
                    _('Row %d: "Dealer Name" is required.') % row_num
                )
                continue

            if not product_name:
                errors.append(
                    _('Row %d: "Product Name" is required (Dealer: "%s").')
                    % (row_num, dealer_name)
                )
                continue

            # ── Dealer lookup ────────────────────────────────────────
            partner_id = partner_cache.get(dealer_name.lower())
            if not partner_id:
                errors.append(
                    _('Row %d: Dealer "%s" not found in Odoo. '
                      'The name must match an existing contact exactly '
                      '(case-insensitive).')
                    % (row_num, dealer_name)
                )
                continue

            # ── Product lookup ───────────────────────────────────────
            product_id = product_cache.get(product_name.lower())
            if not product_id:
                errors.append(
                    _('Row %d: Product "%s" not found in Odoo. '
                      'The name must match an existing product exactly '
                      '(case-insensitive, including variant attributes).')
                    % (row_num, product_name)
                )
                continue

            # ── Qty ──────────────────────────────────────────────────
            qty = self._parse_float(qty_raw)
            if qty is None:
                errors.append(
                    _('Row %d: Invalid Qty "%s" for dealer "%s", '
                      'product "%s". Must be a number (e.g. 10 or 1.5).')
                    % (row_num, qty_raw, dealer_name, product_name)
                )
                continue
            if qty < 0:
                errors.append(
                    _('Row %d: Qty cannot be negative '
                      '(dealer: "%s", product: "%s").')
                    % (row_num, dealer_name, product_name)
                )
                continue

            # ── Rate (optional) ──────────────────────────────────────
            rate = None
            if rate_raw not in (None, ''):
                rate = self._parse_float(rate_raw)
                if rate is None:
                    errors.append(
                        _('Row %d: Invalid Rate "%s" for dealer "%s", '
                          'product "%s". Must be a number, or leave blank '
                          'to use the product default rate.')
                        % (row_num, rate_raw, dealer_name, product_name)
                    )
                    continue

            import_plan.append({
                'row_num': row_num,
                'partner_id': partner_id,
                'product_id': product_id,
                'qty': qty,
                'rate': rate,       # None → resolved from product in _apply
                'dealer_name': dealer_name,
                'product_name': product_name,
            })

        return import_plan, errors

    # ------------------------------------------------------------------
    # Write to dispatch sheet
    # ------------------------------------------------------------------

    def _apply_import_plan(self, import_plan, sheet):
        """
        Write validated rows to the dispatch sheet.

        Algorithm
        ---------
        1. Build an in-memory index of existing dealer lines and their
           product lines so all lookups are O(1).
        2. Group the import plan by partner_id.
        3. For each partner group:
           a. If no dealer line exists → create one via ORM.
           b. For each product in the group:
              - If (dealer_line_id, product_id) already exists → skip + count.
              - Otherwise → create a new product line via ORM.
        4. Return stats dict.

        All creates go through the standard ORM, so every existing
        constraint (_check_unique_partner_per_sheet,
        _check_unique_product_per_line, _check_qty) is fully enforced.
        """
        from collections import defaultdict

        DispatchLine = self.env['milk.dispatch.line']
        ProductLine = self.env['milk.dispatch.product.line']

        # ── Build current-state indexes ───────────────────────────────
        # {partner_id: dispatch_line_record}
        dealer_index = {
            dl.partner_id.id: dl
            for dl in sheet.line_ids
        }
        # {(dispatch_line_id, product_id): True}
        product_index = {
            (pl.dispatch_line_id.id, pl.product_id.id): True
            for dl in sheet.line_ids
            for pl in dl.product_line_ids
        }

        # ── Group import plan by partner ──────────────────────────────
        plan_by_partner = defaultdict(list)
        for entry in import_plan:
            plan_by_partner[entry['partner_id']].append(entry)

        new_dealers = 0
        new_products = 0
        skipped = 0

        for partner_id, entries in plan_by_partner.items():

            # Get or create the dealer line
            if partner_id in dealer_index:
                dispatch_line = dealer_index[partner_id]
            else:
                dispatch_line = DispatchLine.create({
                    'sheet_id': sheet.id,
                    'partner_id': partner_id,
                })
                dealer_index[partner_id] = dispatch_line
                new_dealers += 1
                _logger.debug(
                    'Import: created dispatch line — partner_id=%d sheet=%s',
                    partner_id, sheet.name,
                )

            # Append product lines
            for entry in entries:
                product_id = entry['product_id']
                key = (dispatch_line.id, product_id)

                if key in product_index:
                    skipped += 1
                    _logger.debug(
                        'Import: skipped duplicate — '
                        'partner_id=%d product_id=%d (row %d)',
                        partner_id, product_id, entry['row_num'],
                    )
                    continue

                # Resolve rate: imported value takes priority; blank → product default
                rate = entry['rate']
                if rate is None:
                    rate = self._resolve_product_rate(product_id)

                ProductLine.create({
                    'dispatch_line_id': dispatch_line.id,
                    'product_id': product_id,
                    'qty': entry['qty'],
                    'rate': rate,
                })
                product_index[key] = True
                new_products += 1

        return {
            'new_dealers': new_dealers,
            'new_products': new_products,
            'skipped': skipped,
        }

    # ------------------------------------------------------------------
    # Lookup cache builders
    # ------------------------------------------------------------------

    def _build_partner_cache(self):
        """
        Return {lower_name: partner_id} for ALL res.partner records.

        When two partners share the same name (case-insensitively), the
        record with the lower ID wins — consistent with Odoo's name_search
        default behaviour.
        """
        records = self.env['res.partner'].sudo().search_read(
            [], ['id', 'name'], order='id asc'
        )
        cache = {}
        for rec in records:
            key = (rec['name'] or '').strip().lower()
            if key and key not in cache:
                cache[key] = rec['id']
        return cache

    def _build_product_cache(self):
        """
        Return {lower_display_name: product_id} for all ACTIVE
        product.product records.

        display_name includes variant attribute values (e.g.
        "Full Cream Milk (500ml)") which is what the user sees in
        the Dispatch Sheet product drop-down.
        """
        records = self.env['product.product'].sudo().with_context(
            lang=self.env.user.lang
        ).search_read(
            [('active', '=', True)],
            ['id', 'display_name'],
            order='id asc',
        )
        cache = {}
        for rec in records:
            key = (rec['display_name'] or '').strip().lower()
            if key and key not in cache:
                cache[key] = rec['id']
        return cache

    # ------------------------------------------------------------------
    # Utility
    # ------------------------------------------------------------------

    def _resolve_product_rate(self, product_id):
        """
        Resolve the default billing rate for a product.

        Priority (mirrors MilkDispatchProductLine._onchange_product):
          1. product.template.milk_rate_per_crate  (if > 0)
          2. product.product.lst_price             (standard sales price)
        """
        product = self.env['product.product'].browse(product_id)
        tmpl = product.product_tmpl_id
        if tmpl.milk_rate_per_crate and tmpl.milk_rate_per_crate > 0:
            return tmpl.milk_rate_per_crate
        return product.lst_price

    @staticmethod
    def _to_str(value):
        """
        Safely convert a cell value to a stripped string.
        Returns '' for None, NaN strings, or whitespace-only values.
        """
        if value is None:
            return ''
        s = str(value).strip()
        if s.lower() in ('nan', 'none', 'n/a', '#n/a', ''):
            return ''
        return s

    @staticmethod
    def _parse_float(value):
        """
        Safely parse a cell value as float.
        Returns None if the value is blank or cannot be parsed as a number.
        """
        if value is None:
            return None
        s = str(value).strip()
        if not s or s.lower() in ('nan', 'none', 'n/a', '#n/a'):
            return None
        try:
            return float(s)
        except (ValueError, TypeError):
            return None
